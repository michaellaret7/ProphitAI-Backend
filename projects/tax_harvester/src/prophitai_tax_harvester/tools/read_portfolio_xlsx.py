"""read_portfolio_xlsx — Read the user's tax-lot portfolio workbook.

Parses the three-sheet xlsx the Tax Harvester agent operates on:
    - Lots: every open tax lot (cost, current price, acquisition date)
    - Activity: recent purchases/sales/DRIPs relevant to the wash-sale window
    - YTD Realized: year-to-date realized gain/loss buckets (ST/LT)

Returns the parsed contents plus derived per-lot fields the agent needs
for tax-loss harvesting decisions: market_value, cost_basis,
unrealized_pnl ($ and %), days_held, and holding_character (ST/LT).
"""

from datetime import date, datetime
from pathlib import Path
from typing import Annotated, Any, Optional

from openpyxl import load_workbook

from prophitai_atlas.tools.decorator import agent_tool, Param
from prophitai_atlas.tools.responses import success_response, error_response
from prophitai_shared.time_utils import get_current_utc_time


# ================================
# --> Helper funcs
# ================================

DEFAULT_XLSX_PATH = (
    Path(__file__).resolve().parent.parent / "data" / "sample_portfolio.xlsx"
)

LOTS_SHEET = "Lots"
ACTIVITY_SHEET = "Activity"
YTD_SHEET = "YTD Realized"

LONG_TERM_DAYS = 365


def _read_sheet(workbook, sheet_name: str) -> list[dict[str, Any]]:
    """Read a worksheet into a list of dicts keyed by the header row."""
    if sheet_name not in workbook.sheetnames:
        return []

    ws = workbook[sheet_name]
    rows = ws.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        return []

    headers = [str(h).strip() if h is not None else "" for h in headers]

    parsed: list[dict[str, Any]] = []
    for row in rows:
        if all(cell is None for cell in row):
            continue

        record = {headers[i]: row[i] for i in range(len(headers))}

        parsed.append(record)

    return parsed


def _to_iso_date(value: Any) -> Optional[str]:
    """Convert a date/datetime cell to YYYY-MM-DD; pass through strings; None otherwise."""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, str):
        return value

    return None


def _round(value: Optional[float], ndigits: int = 2) -> Optional[float]:
    if value is None:
        return None

    return round(float(value), ndigits)


def _enrich_lots(
    lots: list[dict[str, Any]],
    today: date,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Attach derived fields to each lot and return per-portfolio aggregates."""
    enriched: list[dict[str, Any]] = []

    total_market_value = 0.0
    total_cost_basis = 0.0
    total_unrealized_pnl = 0.0
    loss_lot_count = 0

    for lot in lots:
        quantity = lot.get("Quantity") or 0
        cost_per_share = lot.get("Cost/Share ($)") or 0
        current_price = lot.get("Current Price ($)") or 0
        acquired_raw = lot.get("Acquisition Date")

        cost_basis = float(quantity) * float(cost_per_share)
        market_value = float(quantity) * float(current_price)
        unrealized_pnl = market_value - cost_basis
        unrealized_pnl_pct = (unrealized_pnl / cost_basis) if cost_basis else None

        acquired_date: Optional[date] = None
        if isinstance(acquired_raw, datetime):
            acquired_date = acquired_raw.date()
        elif isinstance(acquired_raw, date):
            acquired_date = acquired_raw

        days_held = (today - acquired_date).days if acquired_date else None
        character = None
        if days_held is not None:
            character = "long_term" if days_held > LONG_TERM_DAYS else "short_term"

        enriched_lot = {
            "lot_id": lot.get("Lot ID"),
            "symbol": lot.get("Symbol"),
            "description": lot.get("Description"),
            "asset_class": lot.get("Asset Class"),
            "sector": lot.get("Sector"),
            "acquisition_date": _to_iso_date(acquired_raw),
            "quantity": _round(quantity, 4),
            "cost_per_share": _round(cost_per_share, 4),
            "current_price": _round(current_price, 4),
            "cost_basis": _round(cost_basis),
            "market_value": _round(market_value),
            "unrealized_pnl": _round(unrealized_pnl),
            "unrealized_pnl_pct": _round(unrealized_pnl_pct, 4) if unrealized_pnl_pct is not None else None,
            "days_held": days_held,
            "holding_character": character,
        }

        enriched.append(enriched_lot)

        total_market_value += market_value
        total_cost_basis += cost_basis
        total_unrealized_pnl += unrealized_pnl

        if unrealized_pnl < 0:
            loss_lot_count += 1

    aggregates = {
        "lot_count": len(enriched),
        "loss_lot_count": loss_lot_count,
        "total_cost_basis": _round(total_cost_basis),
        "total_market_value": _round(total_market_value),
        "total_unrealized_pnl": _round(total_unrealized_pnl),
        "total_unrealized_pnl_pct": (
            _round(total_unrealized_pnl / total_cost_basis, 4)
            if total_cost_basis else None
        ),
    }

    return enriched, aggregates


def _normalize_activity(activity: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "date": _to_iso_date(row.get("Date")),
            "symbol": row.get("Symbol"),
            "action": row.get("Action"),
            "quantity": row.get("Quantity"),
        }
        for row in activity
    ]


def _normalize_ytd(ytd: list[dict[str, Any]]) -> dict[str, Optional[float]]:
    """Flatten the YTD Realized sheet into a single bucket dict."""
    out: dict[str, Optional[float]] = {
        "st_realized_gain": None,
        "st_realized_loss": None,
        "lt_realized_gain": None,
        "lt_realized_loss": None,
    }

    bucket_map = {
        "ST Realized Gain ($)": "st_realized_gain",
        "ST Realized Loss ($)": "st_realized_loss",
        "LT Realized Gain ($)": "lt_realized_gain",
        "LT Realized Loss ($)": "lt_realized_loss",
    }

    for row in ytd:
        bucket = row.get("Bucket")
        amount = row.get("Amount ($)")

        key = bucket_map.get(str(bucket).strip()) if bucket else None
        if key:
            out[key] = _round(amount)

    return out


# ================================
# --> Tools
# ================================

@agent_tool(name="read_portfolio_xlsx", category="tax_harvester")
def read_portfolio_xlsx(
    file_path: Annotated[
        Optional[str],
        Param(description="Absolute path to the portfolio xlsx. Omit to use the bundled sample_portfolio.xlsx."),
    ] = None,
    losses_only: Annotated[
        bool,
        Param(description="If true, the lots array only contains positions with unrealized loss. Aggregates always span the full portfolio."),
    ] = False,
) -> str:
    """
    Read the user's tax-lot portfolio workbook and return its parsed contents.

    The workbook is the canonical input for tax-loss harvesting. It contains
    three sheets:

    - **Lots**: every open tax lot with Lot ID, Symbol, Description, Asset Class,
      Sector, Acquisition Date, Quantity, Cost/Share, and Current Price. The
      tool enriches each lot with derived fields used by the harvesting logic:
      cost_basis, market_value, unrealized_pnl, unrealized_pnl_pct, days_held,
      and holding_character ("short_term" / "long_term"; long-term is >365
      days held as of today UTC).
    - **Activity**: recent purchases, sales, and dividend reinvestments. Used
      to detect wash-sale-window conflicts in the prior 30 days.
    - **YTD Realized**: year-to-date realized gain/loss buckets split into
      short-term and long-term. Used to plan loss-offset capacity.

    Aggregates across ALL lots (total cost basis, market value, unrealized
    P&L, lot count, loss-lot count) are always returned so the agent does
    not need to recompute them — they are unaffected by ``losses_only``.

    Args:
        file_path: Absolute path to the portfolio xlsx file. Omit (None) to
            use the bundled sample workbook at
            ``prophitai_tax_harvester/data/sample_portfolio.xlsx``. An empty
            string is rejected.
        losses_only: When True, the ``lots`` array only contains positions
            with negative unrealized P&L (the harvest candidate set).
            Aggregates still span the full portfolio. Use this in
            scan-for-candidates mode to halve token cost on large portfolios.

    Returns:
        YAML-formatted result with:
            - file_path: resolved absolute path that was read
            - as_of_date: today's UTC date used for ST/LT classification
            - filter: dict echoing applied filters (e.g., {"losses_only": true})
            - aggregates: portfolio-level totals (always full portfolio)
            - lots: list of enriched per-lot records (filtered if losses_only)
            - activity: list of recent BUY/SELL/DRIP rows
            - ytd_realized: dict of ST/LT realized gain and loss buckets

    Examples:
        read_portfolio_xlsx()
        >>> {"success": True, "data": {"lots": [...all 35 lots...], ...}}

        read_portfolio_xlsx(losses_only=True)
        >>> {"success": True, "data": {"lots": [...17 loss lots...], ...}}

        read_portfolio_xlsx(file_path="/absolute/path/to/portfolio.xlsx")
        >>> {"success": True, "data": {...}}

    Raises:
        Exception: If the workbook cannot be parsed.
    """
    if file_path is not None and not file_path:
        return error_response("file_path must be a non-empty string or omitted; got empty string")

    resolved_path = (Path(file_path) if file_path is not None else DEFAULT_XLSX_PATH).resolve()

    if not resolved_path.is_file():
        return error_response(f"xlsx file not found at: {resolved_path}")

    try:
        workbook = load_workbook(resolved_path, data_only=True, read_only=True)

        raw_lots = _read_sheet(workbook, LOTS_SHEET)
        raw_activity = _read_sheet(workbook, ACTIVITY_SHEET)
        raw_ytd = _read_sheet(workbook, YTD_SHEET)

        workbook.close()

        today = get_current_utc_time().date()

        enriched_lots, aggregates = _enrich_lots(raw_lots, today)
        normalized_activity = _normalize_activity(raw_activity)
        normalized_ytd = _normalize_ytd(raw_ytd)

        lots_payload = (
            [lot for lot in enriched_lots if (lot.get("unrealized_pnl") or 0) < 0]
            if losses_only
            else enriched_lots
        )

        return success_response({
            "file_path": str(resolved_path),
            "as_of_date": today.isoformat(),
            "filter": {"losses_only": losses_only},
            "aggregates": aggregates,
            "lots": lots_payload,
            "activity": normalized_activity,
            "ytd_realized": normalized_ytd,
        })

    except KeyError as e:
        return error_response(f"Missing expected column in workbook: {e}")
    except Exception as e:
        return error_response(f"Error reading portfolio xlsx: {e}")


# ================================
# --> Standalone testing
# ================================

if __name__ == "__main__":
    print("Test 1: Default bundled workbook")
    print(read_portfolio_xlsx())
    print()

    print("Test 2: losses_only=True")
    print(read_portfolio_xlsx(losses_only=True))
    print()

    print("Test 3: Missing file")
    print(read_portfolio_xlsx(file_path="/tmp/does_not_exist.xlsx"))
    print()

    print("Test 4: Directory path")
    print(read_portfolio_xlsx(file_path="/tmp"))
    print()

    print("Test 5: Empty string")
    print(read_portfolio_xlsx(file_path=""))
