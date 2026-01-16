"""
Excel ingestion module for RAG pipelines.

Handles .xlsx and .xls file extraction with support for merged cells,
formulas, and multiple sheets. Outputs plain text or markdown tables.
"""
import logging
from pathlib import Path

from app.core.foundry.models.ingestion_output import Document
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
import xlrd

logger = logging.getLogger(__name__)


class ExcelIngestor:
    """
    A unified Excel ingestion class for RAG pipelines.

    Handles .xlsx and .xls files with support for:
    - Multiple sheets
    - Merged cells (value propagation)
    - Formula evaluation (computed values)
    - Markdown table output for structured data

    Attributes:
        output_format: Output format - "text" (plain) or "markdown" (tables).
        include_sheet_names: Include sheet names as headers in output.
        skip_empty_rows: Skip rows that contain no data.
    """

    def __init__(
        self,
        output_format: str = "text",
        include_sheet_names: bool = True,
        skip_empty_rows: bool = True,
    ) -> None:
        """
        Initialize ExcelIngestor with extraction options.

        Args:
            output_format: "text" for plain text, "markdown" for table format.
            include_sheet_names: Whether to include sheet names as headers.
            skip_empty_rows: Whether to skip rows with no data.
        """
        if output_format not in ("text", "markdown"):
            raise ValueError(f"output_format must be 'text' or 'markdown', got: {output_format}")

        self.output_format = output_format
        self.include_sheet_names = include_sheet_names
        self.skip_empty_rows = skip_empty_rows

    def process(self, file_path: str | Path) -> Document:
        """
        Main entry point to extract text from an Excel file.

        Args:
            file_path: Path to the Excel file (.xlsx or .xls).

        Returns:
            Document with content and metadata.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file format is not supported.
            RuntimeError: If extraction fails.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = path.suffix.lower()
        logger.info(f"Processing Excel file: {path.name}")

        if suffix == ".xlsx":
            content, sheet_count = self._extract_xlsx(path)
        elif suffix == ".xls":
            content, sheet_count = self._extract_xls(path)
        else:
            raise ValueError(f"Unsupported file format: {suffix}. Expected .xlsx or .xls")

        metadata = {
            "filename": path.name,
            "extension": suffix,
            "size_bytes": path.stat().st_size,
            "char_count": len(content),
            "sheet_count": sheet_count,
            "output_format": self.output_format,
        }

        return Document(
            content=content,
            metadata=metadata,
            source=str(path.absolute()),
        )

    def _extract_xlsx(self, file_path: Path) -> tuple[str, int]:
        """
        Extract text from .xlsx files using openpyxl.

        Args:
            file_path: Path to the .xlsx file.

        Returns:
            Tuple of (extracted text content, sheet count).

        Raises:
            RuntimeError: If extraction fails.
        """
        try:
            # Reason: data_only=True returns computed formula values instead of formulas
            wb = load_workbook(file_path, data_only=True)
            sheets_output: list[str] = []
            sheet_count = len(wb.sheetnames)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_text = self._extract_sheet_xlsx(ws, sheet_name)
                if sheet_text.strip():
                    sheets_output.append(sheet_text)

            wb.close()

            result = "\n\n".join(sheets_output)
            logger.debug(f"Extracted {len(result)} chars from {sheet_count} sheets")
            return result, sheet_count

        except Exception as e:
            logger.error(f"XLSX extraction failed: {e}")
            raise RuntimeError(f"Failed to extract XLSX: {e}") from e

    def _extract_sheet_xlsx(self, ws: Worksheet, sheet_name: str) -> str:
        """
        Extract text from a single xlsx worksheet.

        Args:
            ws: The openpyxl Worksheet object.
            sheet_name: Name of the sheet for headers.

        Returns:
            Extracted text from the sheet.
        """
        # Build merged cell map for value propagation
        merged_values = self._build_merged_cell_map(ws)

        rows_data: list[list[str]] = []
        for row in ws.iter_rows():
            row_values: list[str] = []
            for cell in row:
                value = self._get_cell_value(cell, merged_values)
                row_values.append(value)

            # Skip empty rows if configured
            if self.skip_empty_rows and not any(v.strip() for v in row_values):
                continue

            rows_data.append(row_values)

        if not rows_data:
            return ""

        # Format output
        if self.output_format == "markdown":
            return self._format_as_markdown(rows_data, sheet_name)
        else:
            return self._format_as_text(rows_data, sheet_name)

    def _build_merged_cell_map(self, ws: Worksheet) -> dict[str, str]:
        """
        Build a map of merged cell coordinates to their values.

        When cells are merged, only the top-left cell contains the value.
        This map allows retrieving that value for any cell in the merged range.

        Args:
            ws: The openpyxl Worksheet object.

        Returns:
            Dict mapping cell coordinates to merged cell values.
        """
        merged_values: dict[str, str] = {}

        for merged_range in ws.merged_cells.ranges:
            # Get the top-left cell value
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            value = str(top_left.value) if top_left.value is not None else ""

            # Map all cells in the range to this value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    coord = f"{row},{col}"
                    merged_values[coord] = value

        return merged_values

    def _get_cell_value(self, cell, merged_values: dict[str, str]) -> str:
        """
        Get the value of a cell, handling merged cells.

        Args:
            cell: The openpyxl cell object.
            merged_values: Map of merged cell coordinates to values.

        Returns:
            String value of the cell.
        """
        coord = f"{cell.row},{cell.column}"

        # Check if this is a merged cell
        if coord in merged_values:
            return merged_values[coord]

        # Handle MergedCell type (these have no value attribute)
        if isinstance(cell, MergedCell):
            return ""

        # Normal cell
        if cell.value is None:
            return ""
        return str(cell.value)

    def _extract_xls(self, file_path: Path) -> tuple[str, int]:
        """
        Extract text from legacy .xls files using xlrd.

        Args:
            file_path: Path to the .xls file.

        Returns:
            Tuple of (extracted text content, sheet count).

        Raises:
            RuntimeError: If extraction fails.
        """
        try:
            wb = xlrd.open_workbook(str(file_path))
            sheets_output: list[str] = []
            sheet_count = wb.nsheets

            for sheet_idx in range(sheet_count):
                sheet = wb.sheet_by_index(sheet_idx)
                sheet_text = self._extract_sheet_xls(sheet)
                if sheet_text.strip():
                    sheets_output.append(sheet_text)

            result = "\n\n".join(sheets_output)
            logger.debug(f"Extracted {len(result)} chars from {sheet_count} sheets")
            return result, sheet_count

        except Exception as e:
            logger.error(f"XLS extraction failed: {e}")
            raise RuntimeError(f"Failed to extract XLS: {e}") from e

    def _extract_sheet_xls(self, sheet) -> str:
        """
        Extract text from a single xls worksheet.

        Args:
            sheet: The xlrd Sheet object.

        Returns:
            Extracted text from the sheet.
        """
        rows_data: list[list[str]] = []

        for row_idx in range(sheet.nrows):
            row_values: list[str] = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                value = str(cell_value) if cell_value else ""
                row_values.append(value)

            # Skip empty rows if configured
            if self.skip_empty_rows and not any(v.strip() for v in row_values):
                continue

            rows_data.append(row_values)

        if not rows_data:
            return ""

        # Format output
        if self.output_format == "markdown":
            return self._format_as_markdown(rows_data, sheet.name)
        else:
            return self._format_as_text(rows_data, sheet.name)

    def _format_as_text(self, rows: list[list[str]], sheet_name: str) -> str:
        """
        Format extracted rows as plain text.

        Args:
            rows: List of row data (each row is a list of cell values).
            sheet_name: Name of the sheet.

        Returns:
            Formatted plain text.
        """
        output: list[str] = []

        if self.include_sheet_names:
            output.append(f"--- Sheet: {sheet_name} ---")

        for row in rows:
            # Join non-empty cells with spaces
            row_text = " | ".join(v for v in row if v.strip())
            if row_text.strip():
                output.append(row_text)

        return "\n".join(output)

    def _format_as_markdown(self, rows: list[list[str]], sheet_name: str) -> str:
        """
        Format extracted rows as a markdown table.

        Args:
            rows: List of row data (each row is a list of cell values).
            sheet_name: Name of the sheet.

        Returns:
            Formatted markdown table.
        """
        if not rows:
            return ""

        output: list[str] = []

        if self.include_sheet_names:
            output.append(f"## {sheet_name}")
            output.append("")

        # Calculate column widths for alignment
        num_cols = max(len(row) for row in rows)

        # Normalize row lengths
        normalized_rows = [row + [""] * (num_cols - len(row)) for row in rows]

        # First row as header
        header = normalized_rows[0]
        output.append("| " + " | ".join(h if h.strip() else "-" for h in header) + " |")
        output.append("| " + " | ".join("---" for _ in header) + " |")

        # Data rows
        for row in normalized_rows[1:]:
            output.append("| " + " | ".join(v if v.strip() else "-" for v in row) + " |")

        return "\n".join(output)

